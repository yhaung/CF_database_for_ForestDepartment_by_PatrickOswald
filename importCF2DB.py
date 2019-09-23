######
## import processed + checked CF data into OMM CF master database

## prechecks
# folder name does not yet exist in cf_permits  : done
# CF-code does not yet exist in excel and polygon FC : done
# all columns in excel file exist (use column list from master-cf with exception for [d_import])  : done
# all columns in excel file have cf_code  : done

# check for files that can not be copied (long-filenames) #rename those with manual name and confirmation dialog : done


# count and display number of columns to import, display c_cf
#have mandatory python confirmation before running : done


## to-do
# copy folders to cf_permits master folder  :done
# create backup of excel master database + polygon master feature class  :done
# import polygon to cf-master polygon feature class  :done
# import excel rows to CF master excel database
# create joined product polygon and point-cf feature classes

## ------------------ start import libraries --------------------- ##
from xlrd import open_workbook
import openpyxl
from pathlib import Path
import sys
import arcpy
import shutil
import errno
import re
from datetime import datetime
## ------------------ end import libraries --------------------- ##



## ------------------ start define global variables --------------------- ##
#- full path to input excel workbook
xlsBook =r"L:\OMM_projectMaster\mm_cfcertificates\forImport\20190829_CF_permits_nok\20190829_CFpermitsFD_OMM.xlsx"
#- name of excel sheet to import
xlsSheet =r"cfCertificatesDB"

#- full path to master excel workbook (target for import)
xlsBookMaster = r"L:\OMM_projectMaster\mm_cfcertificates\z_master_cfdb_omm\sourcedata\master_mm_cfcertificateFD_omm.xlsx"
#- name of excel sheet to import (target for import)
xlsSheetMaster =r"cfCertificatesDB"
#- name of sheet containing the Schema for the CF DB sheet
xlsSheetSchema = r"CFDBschema"
#- name of sheet containing the configurations for the CF DB sheet
xlsSheetConfig = r"configuration"
#- cell holding the base folder value
xlsCellXYConfigBasefolder = r"B3"

#- full path to shapefile of import dataset cf-boudnaries
shpcf = r"L:\OMM_projectMaster\mm_cfcertificates\forImport\20190829_CF_permits_nok\cf_vectorData\cfCertificates_nwe.shp"
#- full path of CF master geodatabase
gdbcfMaster = r"L:\OMM_projectMaster\mm_cfcertificates\z_master_cfdb_omm\sourcedata\mm_cf_omm.gdb"
#-subpath (inside GDB) to master cf-boundary polygon feature class
fccfMaster = r"master_mm_cf_omm\mm_cfCertificatesFD_omm_py"

#- full path of source cf_permit rootfolder
dir_srcbase = r"L:\OMM_projectMaster\mm_cfcertificates\forImport\20190829_CF_permits_nok\cf_permits"
#- full path of target/master cf_permit rootfolder
dir_trgbase = r"L:\OMM_projectMaster\mm_cfcertificates\z_master_cfdb_omm\sourcedata\cf_permits"

#- full dir of root of cf database
dir_cfbase = "L:\OMM_projectMaster\mm_cfcertificates\z_master_cfdb_omm\sourcedata"

#- full path or folder to store backups of master GDB and Excel file
dir_backup = r"L:\OMM_projectMaster\mm_cfcertificates\z_master_cfdb_omm\archive\backups"


dir_gis = "gis"  #subpath to gis data
dir_scan = "scan"  #subpath to scan data
dir_otherdoc = "other_docs" #subpath to other data
subdirs_mandatory = [dir_gis, dir_scan, dir_otherdoc]

colfc_c_cf = "c_CF" # column name with omm cf-code / unique ID in the feature class and shapefile

col_c_cf = "c_CF"   # column name with omm cf-code / unique ID
col_nm_suffix = "nm_suffix" # column name with cf name suffix (village shorty)
col_doc_cert = "doc_cert" # column name with certificate-pdf status
col_doc_appl = "doc_appl" # column name with application-pdf status
col_doc_fsr = "doc_fsr" # column name with field survey report-pdf status
col_doc_mngtpl = "doc_mngtpl" # column name with management plan-pdf status
col_doc_vfv = "doc_vfv" # column name with VFV approval letter-pdf status
col_certmap = "certmap" # column name with CF boundary map status
col_FDorgShp = "FDorgShp" # column name with FD original/provided CF boundary shapefile status

cols_nonMandatory = ["D_IMPTBL","geomIssues"]


#column for metadata for import datestamp
col_D_IMPTBL = "D_IMPTBL"
col_IMPBY = "IMP_BY"



#- field names in import/shp fiele to be imported
fccol_shp = ["c_CF","GEOMCOM","SUBMIT_BYD"]                 #need to be in this order, DO NOT CHANGE order
#- field names in target/master fc where data gets added to
fccol_gdbfc = ["c_cf","GEOMCOM","CREATEBY","D_IMPORT"]      #need to be in this order, DO NOT CHANGE order
#str_D_IMPORT = '2019/08/31'  # value to add to date of import (could be dynamic, but then.. often want to speficy manually for certain reason
#str_D_IMPORT = datetime.now().strftime("%Y/%m/%d")  #YYYY/MM/DD
str_D_IMPORT = "2019/09/09"  #YYYY/MM/DD
str_IMP_BY = "PO"  #data import performed by

logfile = dir_cfbase + "\\" + "logfile_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".txt"

#regex to select specific records, column name + regular expression for input excel
xlsfilter = ["D_SUBMIT", "NOK_2019-08-29"]  #^MMR1002   \w

# to select specific records in shp, column name + sql expression for input shapefile
shpfilter = ["SUBMIT_BYD", "NOK_2019-08-29"]
shp_sqlwhereexp = """{0} = '{1}'""".format(arcpy.AddFieldDelimiters(shpcf, shpfilter[0]), shpfilter[1])

#- columns that will be filled with dynamically calculated values
formula_cols = ["f_doccert","f_docappl","f_docfsr","f_docmngt","f_docvfv","f_certmap","l_dir","l_doccert","l_docappl","l_docfsr","l_docmngt","l_docvfv","l_certmap"]

# switch if polygones should be imported
switch_polygons2import = False

## ------------------ end define global variables --------------------- ##


## ------------------ start define functions --------------------- ##

## ++++++++++++++ start read excel sheet to dictionary list ++++++++++++++++++ ##
def xls2list(xlsBook, xlsSheet):
    """
    This function read an excel sheet and returns it as a list where each row is a
    dicctionary with the column name as the dictionary-key and the cell value as the dictionary-value
    :param xlsBook: full path to input excel workbook
    :param xlsSheet: name of excel sheet
    :return: list of dictionaries where each excel row is one dictionary
    """
    #- checking if input file exists
    my_file = Path(xlsBook)
    if not my_file.is_file():  # check for valid file path
        print("File",xlsBook, "does not exist. Please check for correct spelling. Script will terminate." )
        sys.exit()
    # using xlrd
    dict_list = []
    book = open_workbook(xlsBook)
    sheet = book.sheet_by_name(xlsSheet)
    #reads 1st row for column names

    #legacy: colNames = sheet.row_values(0)  #storing column names to be able to store them back in excel
    keys = sheet.row_values(0)
    # read the rest rows for values and stores them in the list
    values = [sheet.row_values(i) for i in range(1, sheet.nrows)]
    for value in values:
        dict_list.append(dict(zip(keys, value)))

    if len(dict_list) > 0:
        return dict_list
    else:
        print ("Problem importing data from excel. Script will terminate.")
        sys.exit()
## ++++++++++++++ end read excel sheet to dictionary list ++++++++++++++++++ ##

## ++++++++++++++ start check for valid mandatory columns in import sheet ++++++++++++++++++ ##
def check4validcolumns(srcCols,xlsMasterSchemaKeys):
    """
    tests if columns in import xls sheet are valid (must be equal to those in master/target excel sheet except
    for those columns defined as non-mandatory in a global list and those fields that get generated dynamically)
    :param srcCols: keys / column names in source/import excel sheet
    :param masterSheetSchema: list with all mandatory columns in target/master excel sheet
    :return: true if all mandatory fields are found in import dataset, false if one or more are missing
    """

    # - remove non-mandatory fields from master column list for testing
    log_issues = list()
    xlsMasterKeys_mandatory = list(xlsMasterSchemaKeys)
    print (xlsMasterKeys_mandatory)
    for col_nonMandatory in cols_nonMandatory:
        try:
            xlsMasterKeys_mandatory.remove(col_nonMandatory)
        except:
            print(col_nonMandatory + " not found in schema.")
            pass
    for formula_col in formula_cols:
        try:
            xlsMasterKeys_mandatory.remove(formula_col)
        except:
            print(formula_col + " not found in schema.")
            pass
    print(xlsMasterKeys_mandatory)

    cols_mandatoryMissing = list()
    for xlsMasterKey in xlsMasterKeys_mandatory:
        if xlsMasterKey not in srcCols:
            cols_mandatoryMissing.append(xlsMasterKey)
        # else:
        #     print("Mandatory column", xlsMasterKey, "found in import dataset.")
    if len(cols_mandatoryMissing) > 0:
        for col_mandatoryMissing in cols_mandatoryMissing:
            log_issues.append("Mandatory column" + col_mandatoryMissing + "missing in import excel sheet.")

    if len(log_issues) > 0:
        print("Mandatory columns in master excel sheet vs schema definition issue." )
        for log in log_issues:
            print(log)
        return False
    else:
        return True
## ++++++++++++++ end check for valid mandatory columns in import sheet ++++++++++++++++++ ##


## ++++++++++++++ start checking for valid folders to copy and target location ++++++++++++++++++ ##
def check4validDirs(dataRows,col_cfcode,col_cfsuffix):
    """
    Checks for the existance of source and target directories
    :param dataRows: list dictionary with data from import excel sheet
    :param col_cfcode: cf-code (as it is part of the folder names)
    :param col_cfsuffix: cf-suffix (village sorty) as it is part of the folder names
    :return:  true if no issues, false if issues
    """
    log_dir = list()

    # - check for target root dir (only to be done once as the same for all)
    mydir = Path(dir_trgbase)
    if not mydir.is_dir():
        log_dir.append(dataRows[col_c_cf] + " target root directory missing: " + dir_trgbase)

    for dataRow in dataRows:
        cfcode = dataRow[col_cfcode]
        cfsuffix = dataRow[col_cfsuffix]
        #- check for source dirs
        dirs2check = list() # list with all dirs and subdirs to check
        dirs2check.append(dir_srcbase + "\\" + cfcode + "_" + cfsuffix) #root dir of dataset
            #- prepare list of dirs to check
        for subdir_mandatory in subdirs_mandatory:
            dirs2check.append(dir_srcbase + "\\" + cfcode + "_" + cfsuffix + "\\" + subdir_mandatory)  # sub-dirs to check

        for dir2check in dirs2check:
            mydir = Path(dir2check)
            if not mydir.is_dir(): #dir should exist
                log_dir.append(dataRow[col_c_cf] + " missing source directory: " + dir2check)

        #- check for target sub-dirs
        targetdir = (dir_trgbase + "\\" + cfcode + "_" + cfsuffix)
        mydir = Path(targetdir)
        if mydir.is_dir():  # dir should not exist
            log_dir.append(dataRow[col_c_cf] + " target directory already exsits: " + targetdir)

    if len(log_dir) > 0:
        print("Valid directory issue(s): " )
        for log in log_dir:
            print(log)
        return False
    else:
        return True
## ++++++++++++++ end checking for valid folders to copy and target location ++++++++++++++++++ ##

## ++++++++++++++ start checking for dupplicated data between import data and data in featrue class and excel master sheet ++++++++++++++++++ ##
def check4dupplicatedcfCodes(shpImport,fcMaster,dataRowsImport,dataRowsMaster):
    """
    Checks for dupplicated cf-codes in the import datasets vs the target/master datasets.
    If the cf-code from the import datasets is already present in the target/master dataset then there is a problem
    and the data shall not be imported as this would result in dupplicated data (or the cf-codes need to be fixed first)
    :param shpImport: full path of the shp-file with the CF boundary geometry for import
    :param fcMaster: full path of the fc with the CF boundary geometry in the target/master geodatabase
    :param dataRowsImport: list dictionary with data from import excel sheet
    :param dataRowsMaster: list dictionary with data of the target/master excel sheet
    :return: true if no issues were encountered. False is there where some dupplicated cf-coode values.
    """
    log_cfcodeissue = list()
    codes_cf_shp = list()
    codes_cf_fc = list()
    codes_cf_xls = list()
    codes_cf_xlsMaster = list()

    cursor = arcpy.da.SearchCursor(shpImport, [colfc_c_cf, shpfilter[0]], shp_sqlwhereexp)
    for row in cursor:
        # if bool(re.search(shpfilter[1], row[1])):       #uses regualr expression to limit/import only those matching with filter
        codes_cf_shp.append(row[0])
    del cursor
    print(len(codes_cf_shp), "Rows for from shapefile read for processing.")
    #- test if anything has been imported from shapefile
    if len(codes_cf_shp) > 0:
        print(str(len(codes_cf_shp)) + " rows from import shapefile read for further processing.")
        global switch_polygons2import
        switch_polygons2import = True
        cursor = arcpy.da.SearchCursor(fcMaster, [colfc_c_cf])
        for row in cursor:
            codes_cf_fc.append(row[0])
        del cursor
        for dataRowImport in dataRowsImport:
            codes_cf_xls.append(dataRowImport[col_c_cf])
        for dataRowMaster in dataRowsMaster:
            codes_cf_xlsMaster.append(dataRowMaster[col_c_cf])

        #- check for dupplicate in geodata
        for code_cf_shp in codes_cf_shp:
            if code_cf_shp in codes_cf_fc:
                log_cfcodeissue.append(code_cf_shp + " already exist in master GDB.")
        #- check for dupplicates in excel sheets
        for code_cf_xls in codes_cf_xls:
            if code_cf_shp in codes_cf_xlsMaster:
                log_cfcodeissue.append(code_cf_shp + " already exist in master Excel sheet.")
        # - check for missing entry in input excel vs. shapefile (should already have been checked before but still cant harm to double check
        for code_cf_shp in codes_cf_shp:
            if code_cf_shp not in codes_cf_xls:
                log_cfcodeissue.append(code_cf_shp + " has no matching pair in import excel sheet.")
    else:
        print("No datarows from import shapefile read. Please check your shapefile filter definition.")
        print("Your current filter definition is " + shp_sqlwhereexp)
        print("No Geometry will be imported!!!")

        #- usually there should be at least some polygons to be imported. But it can be that there are only
        #  data for the excel file and folders if there were no suitable maps for any of the CFs.
        proceed = ""
        while proceed not in ('y', 'n'):
            if proceed != "":
                print("Please type 'y' to continue or 'n' to abort. (followed by ENTER key)")
            proceed = str(input("No polygons found in shapefile for import. Continue with data imports (y/n)?")).lower()
            if proceed == 'n':
                print("User abort processing. Data will not be imported. Script will terminate.")
                sys.exit()
            elif proceed == 'y':
                print("Data import without importing any polygons chosen. Script will continue to run.")
                break  # exits while loop


    if len(log_cfcodeissue) > 0:
        print("CF code / dupplicated datasets issue(s): ")
        for log in log_cfcodeissue:
            print(log)
        return False
    else:
        return True
## ++++++++++++++ end checking for dupplicated data between import data and data in featrue class and excel master sheet ++++++++++++++++++ ##



## ++++++++++++++ start copy import data folders/files to master cf_permit folder ++++++++++++++++++ ##
def copyFolders(dataRows):
    """
    Copies complete folders (and subfolders) from a source to a destination based on the excel sheet list. Fixed for long lath names.
    :param dataRows: list dictionary with data from import excel sheet
    :return: True if all files were copied, False if there were any files that failed to copy.
    """
    log_success = list()
    log_copyfileissues = list()
    filecounter = 0
    for dataRow in dataRows:
        src = dir_srcbase + "\\" + dataRow[col_c_cf] + "_" + dataRow[col_nm_suffix]
        dst = dir_trgbase + "\\" + dataRow[col_c_cf] + "_" + dataRow[col_nm_suffix]
        try:
            shutil.copytree("\\\\?\\" + src, "\\\\?\\" + dst)  # copies folders  "\\\\?\\"  is to allow long path names
            filecounter = filecounter + 1  # to count of copied files
        except OSError as e:
            # If the error was caused because the source wasn't a directory
            if e.errno == errno.ENOTDIR:
                shutil.copy("\\\\?\\" + src, "\\\\?\\" + dst)  # copies files "\\\\?\\"  is to allow long path names
                filecounter = filecounter + 1  # to count of copied files
            else:
                log_copyfileissues.append('Data copy error: Error: %s' % e)
                print('Directory or file not copied. Error: %s' % e)
        log_success.append('Copy files for ' + dataRow[col_c_cf] + ' finished.')
        if len(log_copyfileissues) > 0:
            log_success.append('The following files or folders were NOT copied for ' + dataRow[col_c_cf] + ': ')
            for log_copyfileissue in log_copyfileissues:
                log_success.append(log_copyfileissue)
    add2log(log_success)
    if len(log_copyfileissues) > 0:
        print(str(len(log_copyfileissues)) + " error(s) in copying files: ")
        for log in log_copyfileissues:
            print(log)
        return False
    else:
        print("CF package Files/Folders successfully copied.")
        return True

## ++++++++++++++ end copy import data folders/files to master cf_permit folder ++++++++++++++++++ ##

## +++++++++++++++ start master schema +++++++++++++++++++++++++++++++++ ##
def readMasterSchema(xlsMaster,sheetSchema):
    """
    reads the first column from the CF excel schema table and creates an dictionary with the column name and position
    :param xlsMaster: excel book that contains the CF master schema sheet
    :param sheetSchema: sheet with the CF master schema
    :return: a dicctionary with key=column name and value = column position
    """
    log_issues = list()
    my_file = Path(xlsMaster)
    if not my_file.is_file():  # check for valid file path
        print("File",xlsBook, "does not exist. Please check for correct spelling. Script will terminate." )
        log_issues.append("Schema in " + xlsBook + "#" + sheetSchema + " does not exist.")
        sys.exit()
    book = openpyxl.load_workbook(xlsMaster)
    sheet = book[sheetSchema]
    columnposition = {}
    indexpos = 0
    for row in range(2, sheet.max_row + 1):
        indexpos += 1
        if len(str(sheet['B' + str(row)].value)) > 0:
            columnposition[str(sheet['A' + str(row)].value)] = indexpos
        else:
            log_issues.append("Empty row in xls database schema at row "  + str(indexpos) )
            print("Empty row in xls database schema at row "  + str(indexpos))
            print("This issue needs to be fixed first. Script will terminate.")
            sys.exit()
    if len(log_issues) > 0:
        print(str(len(log_issues)) + " error(s) in reading Excel database schema: ")
        for log in log_issues:
            print(log)
        return False
    else:
        return columnposition

## +++++++++++++++ end master schema ++++++++++++++++++++++++++++++++ ##

## ++++++++++++++ start make backup of master files ++++++++++++++++++ ##
def makeBackupMaster(dirBackup):
    """
    copies the master GDB and excel file to the backup-folder and adds the current date-time as a prefix
    :param dirBackup: full path of target/backup directory
    :return: returns True if no issues otherwise returns False
    """
    log_issues = list()
    now = datetime.now()
    dt_string = now.strftime("%Y%m%d-%H%M%S")  #formats current date-time to YYYYMMDD-HHMM
    srcgdb = gdbcfMaster
    regexstr = r"\w*\.gdb$"  #finds the name of the gdb from the full path
    trggdb = dirBackup + "\\" + dt_string + "_" + re.findall(regexstr,gdbcfMaster)[0]  #will produce an error if this gdb does not exist
    srcxls = xlsBookMaster
    regexstr = r"\w*\.xlsx$" #finds the name of the excel file from the full path
    trgxls = dirBackup + "\\" + dt_string + "_" + re.findall(regexstr,xlsBookMaster)[0]  #will produce an error if this gdb does not exist
    try:
        ignored = ['*.lock']  #pattern to ignore while copying files (.lock is a gdb lock file and can/should not be copied)
        shutil.copytree(srcgdb, trggdb, ignore=shutil.ignore_patterns(*ignored))  # copies renamed geodatabase folder (incl. files) to backup folder
        add2log(["Backup of Master GDB created at" + trggdb])
    except OSError as e:
        #print('GDB master not copied to backup. Error: %s' % e)
        log_issues.append('GDB master not copied to backup. Error: %s' % e)
    try:
        shutil.copy(srcxls, trgxls)  # copies renamed excel master file to backup
        add2log(["Backup of Master CF Excel file created at" + trgxls])
    except OSError as e:
        #print('Excel master file not copied to backup. Error: %s' % e)
        log_issues.append('GDB master not copied to backup. Error: %s' % e)

    if len(log_issues) > 0:
        print(str(len(log_issues)) + " error(s) in copying files: ")
        for log in log_issues:
            print(log)
        return False
    else:
        return True

## ++++++++++++++ end make backup of master files ++++++++++++++++++ ##

## ++++++++++++++ start import CF boundary polygons into GDB ++++++++++++++++++ ##
def importGeom2GDB():
    """
    Copies polygons from import shape file to master GDB feature class
    :return: True if no issues, otherwise false
    """
    log_success = list()
    log_issues = list()
    fcMaster = gdbcfMaster + "\\" + fccfMaster
    try:
        fieldsTrg = ["SHAPE@",*fccol_gdbfc]  #to add the geometry to the field list at the start
        trgcursor = arcpy.da.InsertCursor(fcMaster, fieldsTrg)
        fieldsSrc = ["SHAPE@",*fccol_shp]
        with arcpy.da.SearchCursor(shpcf, fieldsSrc, shp_sqlwhereexp) as srccursor:
            for row in srccursor:
                trgcursor.insertRow((row[0], row[1], row[2], row[3], str_D_IMPORT))  # problem: i dont know how to assign values by column name rather than list position
                log_success.append("Geometry for " + row[1] + " added to " + fcMaster)
        del trgcursor
        del srccursor
        add2log(log_success)
    except:
        print("Error with data access cursors. ")
        log_issues.append("Data insert into GDB issue.")
        sys.exit()

    if len(log_issues) > 0:
        print(str(len(log_issues)) + " error(s) in copying polygons to master GDB: ")
        for log in log_issues:
            print(log)
        return False
    else:
        return True
## ++++++++++++++ end import CF boundary polygons into GDB ++++++++++++++++++ ##




def check4dataschemavsxlsmastersheet(xlsDBschema, masterKeys):
    """
    performs a check if the excel master sheet conatains all the mandatory fields defined in the schema-sheet/table
    :param xlsDBschema: dictionary with key = column name and value = position
    :param masterKeys: list with all columns present in the CF excel master sheet
    :return: true if all fields are present, false if fields are missing
    """
    log_issues = list()
    #check if DB schema is the same as masterSheet schema
    if not len(list(set(xlsDBschema.keys()) & set(masterKeys))) == len(xlsDBschema.keys()):
        print("Columns in Excel Database schema and Master Excel sheet are different. The following columns are missing in the master sheet:")
        print(list(set(xlsDBschema.keys()) - set(masterKeys)))
        log_issues.append("Columns in Excel Database schema and Master Excel sheet are different.")

    if len(log_issues) > 0:
        print(str(len(log_issues)) + " issues comparing the CFDB schema with the Master sheet: ")
        for log in log_issues:
            print(log)
        print("Please correct. Script will terminate.")
        sys.exit()
        return False
    else:
        return True
## ++++++++++++++ start adding import datarows to excel master sheet ++++++++++++++++++ ##


def importRows2xls(dataRows2add, dataRowsTrg, xlsMasterBook, xlsMasterSheet, xlsDBschema, importKeys, masterKeys):
    """
    creates a date-timestamp prefix renamed backup copy of the excel master sheet in the excel file
    replaces the previous CF excel master sheet with a new one that contains all the mandatory fields of the existing
    CF excel master sheet as well as the values imported from the new/to-import excel sheet
    :param dataRows2add:
    :param dataRowsTrg:
    :param xlsMasterBook:
    :param xlsMasterSheet:
    :param xlsDBschema:
    :param importKeys:
    :param masterKeys:
    :return:
    """
    ##https://automatetheboringstuff.com/chapter12/  help on openpyxl

    log_success = list()
    log_issues = list()
    book = openpyxl.load_workbook(xlsMasterBook)
    #- rename MasterSheet with adding date-time stamp as prefix
    sheetMaster = book[xlsMasterSheet]
    newSheetname = datetime.now().strftime("%Y%m%d-%H%M%S") + xlsMasterSheet
    sheetMaster.title =  newSheetname
    book.save(xlsMasterBook)
    book.close()
    add2log(['Backup of CF Excel sheet created with name: "' + newSheetname + '" in ' + xlsMasterBook])

    #- import datarows into master sheet
        #- prepares all rows to be imported
    dataRowsAll = dataRowsTrg.copy()  # fill with all values already exisiting in Master sheet

    for dataRow2add in dataRows2add:
        dataRows2add = {}
        for masterKey in masterKeys:
            if masterKey in importKeys:
                dataRows2add[masterKey] = dataRow2add[masterKey]
            elif masterKey not in (formula_cols):
                print("Column " + masterKey + " not present in import sheet")
        dataRowsAll.append(dataRows2add)
        #- adds new sheet for the output

    book = openpyxl.load_workbook(xlsMasterBook)
    book.create_sheet(index=0, title=xlsMasterSheet)
    book.save(xlsMasterBook)
    book.close()
    add2log(['New empty excel sheet "' + xlsMasterSheet + '" created in ' + xlsMasterBook + '.'])
        # - writes all values to new sheet

    book = openpyxl.load_workbook(xlsMasterBook)
    trgSheet = book[xlsMasterSheet]
    #write header
    for colName, colPosition in xlsDBschema.items():
        cell2write = trgSheet.cell(row=1, column=colPosition)
        cell2write.value = colName
    for i, row in enumerate(dataRowsAll):
        for colName, colPosition in xlsDBschema.items():
            try:
                # print("row=", i + 2, "column=", colPosition, "value=", row[colName])
                cell2write = trgSheet.cell(row=(i + 2), column=colPosition)
                if colName in formula_cols:
                    newValue = addFormulas2datarows(i + 2, row, colName, xlsDBschema)  #row[colName]
                    cell2write.value = newValue
                    # log_success.append('Value added for ' + row[col_c_cf] + " - column: " + colName + " value:" + newValue)
                else:
                    cell2write.value = row[colName]
                    # log_success.append('Value added for ' + row[col_c_cf] + " - column: " + colName + " value:" + row[colName])
                log_success.append("Values added for: " + row[col_c_cf] + " to " + xlsMasterSheet)
            except:
                print(row[col_c_cf] + " colName: " + colName + " not present in datarow " + str(i +2))
        #     # old xlrt code
        #     if col in col_dateStyle:
        #     sheet.write(i+1, j, row[col],date_format)
        # else:
        #         trgSheet.write(i+1, j, row[col])
        #     cell2write = trgSheet.cell(row=(i + 2), column=j+1)
        #     cell2write.value = row[col]
    trgSheet.freeze_panes = 'A2'
    log_success.append("Top row fixed in: " + xlsMasterSheet)
    book.save(xlsMasterBook)
    book.close()
    log_success.append(str(len(dataRowsAll)) + " (all) data rows added to: " + xlsMasterSheet + " in " + xlsMasterBook)
    add2log(sorted(set(log_success), key=lambda x: log_success.index(x)))   #removes dupplicated while keeping the order


    #- write basefolder to config sheet
    book = openpyxl.load_workbook(xlsMasterBook)
    configSheet = book[xlsSheetConfig]
    configSheet[xlsCellXYConfigBasefolder].value = dir_trgbase
    book.save(xlsMasterBook)
    book.close()
    add2log(["New root-folder '" + dir_trgbase + "' added to Sheet " + xlsSheetConfig])
    print("The CF excel database now contains:" + str(len(dataRowsAll)) + " data rows." )
## ++++++++++++++ end start adding import datarows to excel master sheet ++++++++++++++++++ ##

## ++++++++++++++ start calculate dyanmic cell values ++++++++++++++++++ ##
def addFormulas2datarows(xlsrow, dataRow, columnname, xlsDBschema):
    """
    adds values to fields that are calculated dynamically (file names pf pdfs and jpgs and hyperlinks to those files)
    :param xlsrow: row number in the target excel file for this data set
    :param dataRow: data row for which to calculate the values (dictionary)
    :param columnname: name of the column for which to calculate value/formula
    :param xlsDBschema: dictionary with column names and position in target excel sheet
    :return: value that gets filled to the target cell
    """
    cellxy_c_cf = (openpyxl.utils.get_column_letter(xlsDBschema[col_c_cf])+ str(xlsrow))
    cellxy_c_suffix = (openpyxl.utils.get_column_letter(xlsDBschema[col_nm_suffix]) + str(xlsrow))
    cellxy_basefolder = "configuration!$B$3"
    cellvalue = ""
    try:
        if columnname == "f_doccert" and dataRow[col_doc_cert] not in ['no','na']:
            cellvalue = '=CONCATENATE(' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_cert.pdf")'
        elif columnname == "f_docappl" and dataRow[col_doc_appl] not in ['no','na']:
            cellvalue = '=CONCATENATE(' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_appl.pdf")'
        elif columnname == "f_docfsr" and dataRow[col_doc_fsr] not in ['no','na']:
            cellvalue = '=CONCATENATE(' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_fsr.pdf")'
        elif columnname == "f_docmngt" and dataRow[col_doc_mngtpl] not in ['no','na']:
            cellvalue = '=CONCATENATE(' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_mngtpl.pdf")'
        elif columnname == "f_docvfv" and dataRow[col_doc_vfv] not in ['no','na']:
            cellvalue = '=CONCATENATE(' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_vfv.pdf")'
        elif columnname == "f_certmap" and dataRow[col_certmap] not in ['no','na']:
            cellvalue = '=CONCATENATE(' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',".jpg")'
        elif columnname == "l_dir":
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix +'))'
        elif columnname == "l_doccert" and dataRow[col_doc_cert] not in ['no','na']:
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"\\' + dir_scan + '\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_cert.pdf"))'
        elif columnname == "l_docappl" and dataRow[col_doc_appl] not in ['no','na']:
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"\\' + dir_scan + '\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_appl.pdf"))'
        elif columnname == "l_docfsr" and dataRow[col_doc_fsr] not in ['no','na']:
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"\\' + dir_scan + '\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_fsr.pdf"))'
        elif columnname == "l_docmngt" and dataRow[col_doc_mngtpl] not in ['no','na']:
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"\\' + dir_scan + '\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_mngtpl.pdf"))'
        elif columnname == "l_docvfv" and dataRow[col_doc_vfv] not in ['no','na']:
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"\\' + dir_scan + '\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"_vfv.pdf"))'
        elif columnname == "l_certmap" and dataRow[col_certmap] not in ['no','na']:
            cellvalue = '=HYPERLINK(CONCATENATE(' + cellxy_basefolder + ',"\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',"\\' + dir_gis + '\\",' + cellxy_c_cf + ',"_",' + cellxy_c_suffix + ',".jpg"))'
        else:
            # print("no document for " + dataRow[col_c_cf] + ": " + columnname + " available." )
            cellvalue = ""
        # print(cellvalue)
        return cellvalue
    except:
        print("There is an issue with some excel formula calcuation.")
        return cellvalue
## ++++++++++++++ end calculate dyanmic cell values ++++++++++++++++++ ##

## ++++++++++++++ start function to dictionary list ++++++++++++++++++ ##
## ++++++++++++++ end function sheet to dictionary list ++++++++++++++++++ ##
## ------------------  end define functions --------------------- ##

def add2log(loglist):
    try:
        f = open(logfile, "a+")
    except:
        print("Can not open or create logfile. Script will terminate.")
        sys.exit()
    for logtext in loglist:
        f.write(logtext + "\n")
    f.close()


## ------------------  start run script --------------------- ##
def main():
    #xlsRows = list()       # list dictionary with valid data rows (id <> empty) of import excel sheet
    #xlsRowsMaster = list()   # list dictionary with valid data rows (id <> empty) of CF master excel sheet
    log_prechecks = list()

    #- read xls sheet to list dictionary
    xlsRowsRaw = xls2list(xlsBook, xlsSheet)
    #- remove rows from list where cf-code is empty (e.g. empty xls rows) and selects only those that match the regex
    xlsRows = [x for x in xlsRowsRaw if not len(x[col_c_cf].strip()) <= 0 and bool(re.search(xlsfilter[1], x[xlsfilter[0]]))]
    #- add field for metadata (import when and by whom)
    for xlsRow in xlsRows:
        xlsRow[col_D_IMPTBL] = str_D_IMPORT
        xlsRow[col_IMPBY] = str_IMP_BY
    xlsKeys = list(xlsRows[0])  #list with all column names
    print (len(xlsRows), "rows from excel imported for processing.")

    xlsMasterRowsRaw = xls2list(xlsBookMaster, xlsSheetMaster)
    xlsMasterRows = [x for x in xlsMasterRowsRaw if not len(x[col_c_cf].strip()) <= 0]
    xlsMasterKeys = list(xlsMasterRows[0]) #list with all column names
    print(len(xlsMasterRows), "rows from master excel sheet imported for processing.")

    xlsDBschema = readMasterSchema(xlsBookMaster, xlsSheetSchema)

    #- check if mandatory columns are present in source/import dataset
    if check4validcolumns(xlsKeys, xlsDBschema.keys()):
        print("Check for valid columns successful.")
    else:
        print("Check for valid columns failed.")
        log_prechecks.append("Mandatory column check failed.")
        #- need to finish as otherwise following functions might fail as they might need some mandatory fields
        print("Problem importing data from excel. Script will terminate.")
        # sys.exit()
    # - check for valid directories for source and target
    if check4validDirs(xlsRows,col_c_cf,col_nm_suffix):
        print("Check for valid directories successful.")
    else:
        print("Check for valid directories failed.")
        log_prechecks.append("Valid directory check failed.")
        # - need to finish as otherwise following functions might fail as they might need some mandatory fields
        print("Problem with directories. Script will terminate.")
        # sys.exit()

    # - check for dupplicated cf-codes/datasets
    if check4dupplicatedcfCodes(shpcf, gdbcfMaster + "\\" + fccfMaster, xlsRows, xlsMasterRows):
        print("Check for dupplicated cf-codes in the source and target geodata and the excel sheets successful.")
    else:
        print("Check for dupplicated cf-codes failed.")
        log_prechecks.append("Dupplicated cf-codes check failed.")
        # - need to finish as otherwise following functions might fail as they might need some mandatory fields
        print("Problem with dupplicated cf-codes between the source and the target datasets. Script will terminate.")
        # sys.exit()
    if check4dataschemavsxlsmastersheet(xlsDBschema, xlsMasterKeys):
        print("All columns from CFDB Schema present in Excel Master sheet.")
    else:
        print("Check for consitency of dataschema between DB schema in lut" + xlsSheetSchema + " and in MasterDB Sheet" + xlsSheetMaster + " failed.")
        log_prechecks.append("Excel sheet data schema check failed.")
        # - need to finish as otherwise following functions might fail as they might need some mandatory fields
        print("Problem with data schema consistency. Script will terminate.")
        # sys.exit()


    #- summary for prechecks and user option to continue or abort importing.
    if len(log_prechecks) > 0:
        print("The following critical issue(s) were encoutered during precheck:")
        for logentry in log_prechecks:
            print(logentry)
        print("Problems with the import data. Script will terminate.")
        sys.exit()
    else:
        proceed = ""
        print("Shapefile import setting is " + str(switch_polygons2import))
        while proceed not in ('y','n'):
            if proceed != "":
                print("Please type 'y' to continue or 'n' to abort. (followed by ENTER key)")
            proceed = str(input("Prechecks successful. Continue with data imports (y/n)?")).lower()
            if proceed == 'n':
                print("User abort processing. Data will not be imported. Script will terminate.")
                sys.exit()
            elif proceed == 'y':
                print("Data import chosen. Script will continue and try to import data.")
                break   #exits while loop

    print("Start importing.")

    # make backup of master GDB and Excel file
    if makeBackupMaster(dir_backup):
        print("Created backup of master GDB and master Excel-file in " + dir_backup)
    else:
        print("Not all files from the Master GDB or Excel-file were copied to the backup-directory. Please check the logfiles.")
        print("This might be a critical problem. Script will terminate.")
        sys.exit()

    # import excel data to master CF excel table
    importRows2xls(xlsRows, xlsMasterRows, xlsBookMaster, xlsSheetMaster, xlsDBschema, xlsKeys, xlsMasterKeys)

    # import CF boundary polygons into GDB
    if switch_polygons2import == True:
        if importGeom2GDB():
            print("Importing polygons into Master GDB feature class " + gdbcfMaster + "\\" + fccfMaster + " successful.")
        else:
            print("Not all polygons from the import shapefile were copied to the master GDB. Please check the logfiles.")
            print("This might be a critical problem. Script will terminate.")
            sys.exit()
    else:
        print("No polygons were imported to GDB as there was either an issue with checking for dupplicated CF codes or no polygons were in the import dataset.")

    #copy files and folders to target directory
    if copyFolders(xlsRows):
        print("All filed and folders successfully copied to destination.")
    else:
        print("Not all files or folders were copied to destination. Please check the logfiles.")
        sys.exit()


    # for xlsRow in xlsRows:
    #     print (xlsRow)
    print("Script finished running successfully.")

## ------------------  end run script --------------------- ##



if __name__ == "__main__":
    main()